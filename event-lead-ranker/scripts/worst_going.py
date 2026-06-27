#!/usr/bin/env python3
"""Rank the GOING pool worst-first for review/bumping. EXCLUDES sponsor staff (never bump co-hosts' team),
KEEPS multipliers, flags bump candidates (students/competitors/inflated). Email in col A for easy copy.
Usage: python worst_going.py --config config.json [--dir .] [--n 40]"""
import json, os, argparse, collections
ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True); ap.add_argument("--dir", default="."); ap.add_argument("--n", type=int, default=40)
A = ap.parse_args()
C = json.load(open(A.config)); SP = [s["name"] for s in C["sponsors"]]
REC = json.load(open(os.path.join(A.dir, "leads-ranked.json")))
going = [r for r in REC if r["status"] == "going" and r["type"] != "Employee"]
for r in going: r["total"] = sum(int(r[s]) for s in SP)
going.sort(key=lambda r: (r["total"], r["name"].lower()))

def reco(r):
    if r["multiplier"]: return "KEEP — multiplier/amplifier"
    if r.get("competitor_of"): return f"BUMP? competitor ({r['competitor_of']})"
    if "Student" in r["type"]: return "BUMP? student"
    vf = r.get("vflag", "")
    if vf in ("fake-company", "micro-company", "wrong-company", "wrong-bigco"): return f"BUMP? {vf}"
    if vf in ("inflated-title", "unconfirmed"): return f"REVIEW — {vf}"
    if r["total"] <= 16: return "REVIEW — low fit"
    return "borderline"

worst = going[:A.n]
cols = ["Email", "Name", "Company", "Title", "Total"] + SP + ["Type", "Verified", "Flag", "Recommendation"]
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Worst going (review)"; ws.append(cols)
    for r in worst:
        ws.append([r["email"], r["name"], r["company"], r["title"], r["total"]] + [r[s] for s in SP] +
                  [r["type"], r.get("verified", ""), r.get("vflag", ""), reco(r)])
    for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=C["ui"]["accent"])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 22
    ws.column_dimensions[get_column_letter(len(cols))].width = 30
    for i, r in enumerate(worst, 2):
        rc = reco(r)
        fill = PatternFill("solid", fgColor="D8F5E0") if rc.startswith("KEEP") else PatternFill("solid", fgColor="FFD6E0") if rc.startswith("BUMP") else None
        if fill: ws.cell(i, len(cols)).fill = fill
    wb.save(os.path.join(A.dir, "worst-going-review.xlsx"))
    print(f"wrote worst-going-review.xlsx ({len(worst)} rows, excludes sponsor staff)")
except ImportError:
    import csv
    with open(os.path.join(A.dir, "worst-going-review.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(cols)
        for r in worst: w.writerow([r["email"], r["name"], r["company"], r["title"], r["total"]] + [r[s] for s in SP] + [r["type"], r.get("verified", ""), r.get("vflag", ""), reco(r)])
    print("wrote worst-going-review.csv")
print("reco tally:", dict(collections.Counter(reco(r) for r in worst)))
