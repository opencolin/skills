#!/usr/bin/env python3
"""From leads-ranked.json build: a self-contained HTML dashboard (search, per-sponsor sort, filter chips,
HubSpot/Salesforce CRM export), TOP-LEADS.md (Top-N per sponsor + multipliers), and a per-sponsor Top-N xlsx.
N sponsors, all driven by config. Usage: python build_dashboard.py --config config.json [--dir .] [--top 25]"""
import json, os, argparse
ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True); ap.add_argument("--dir", default="."); ap.add_argument("--top", type=int, default=25)
A = ap.parse_args()
C = json.load(open(A.config)); SP = [s["name"] for s in C["sponsors"]]
REC = json.load(open(os.path.join(A.dir, "leads-ranked.json")))
ACCENT = C["ui"]["accent"]

def eligible(r, s): return r["type"] != "Employee" and s not in (r.get("competitor_of") or "") and "Investor" not in r["type"]
vmark = lambda r: {"V": "✓", "P": "~", "U": "?"}.get(r.get("verified", ""), "")

# ---------- TOP-LEADS.md ----------
md = ["# Top leads to personally invite — per sponsor", "",
      f"From {len(REC)} engaged guests. Excludes sponsor staff, that sponsor's competitors, and investors. ✓=web-verified.", ""]
for s in SP:
    elig = sorted([r for r in REC if eligible(r, s)], key=lambda r: -int(r[s]))[:A.top]
    md += [f"## {s} — top {A.top}", "", "| # | Score | ✓ | Name | Title | Company | Notes |", "|--|--|--|--|--|--|--|"]
    for i, r in enumerate(elig, 1):
        notes = " / ".join(x for x in [r.get("vflag", ""), r.get("notable", "")] if x)
        md.append(f"| {i} | {r[s]} | {vmark(r)} | {r['name']}{' 📣' if r['multiplier'] else ''} | {r['title']} | {r['company']} | {notes} |")
    md.append("")
order = {"high": 0, "med": 1, "low": 2, "na": 3, "": 4}
mult = sorted([r for r in REC if r["multiplier"]], key=lambda r: (order.get(r.get("influence", ""), 5), r["name"].lower()))
md += ["## 📣 Multipliers / amplifiers (by verified influence — personal-invite list)", "",
       "| Influence | ✓ | Name | Role @ Company | Notable |", "|--|--|--|--|--|"]
for r in mult:
    md.append(f"| **{r.get('influence','') or '?'}** | {vmark(r)} | {r['name']} | {r['title']} @ {r['company']} | {r.get('notable','')} |")
open(os.path.join(A.dir, "TOP-LEADS.md"), "w").write("\n".join(md))

# ---------- per-sponsor Top-N xlsx ----------
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    cols = ["#", "Score", "Name", "Title", "Company", "Type", "Verified", "Influence", "Notable", "Flag", "LinkedIn", "Email"]
    for s in SP:
        ws = wb.create_sheet(s[:31]); ws.append(cols)
        top = sorted([r for r in REC if eligible(r, s)], key=lambda r: -int(r[s]))[:A.top]
        for i, r in enumerate(top, 1):
            ws.append([i, r[s], r["name"], r["title"], r["company"], r["type"], r.get("verified", ""), r.get("influence", ""), r.get("notable", ""), r.get("vflag", ""), r.get("linkedin", ""), r["email"]])
        for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=ACCENT)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        ws.conditional_formatting.add(f"B2:B{A.top+1}", ColorScaleRule(start_type="num", start_value=1, start_color="F8696B", mid_type="num", mid_value=5, mid_color="FFEB84", end_type="num", end_value=9, end_color="63BE7B"))
        for col, w in {"A": 4, "C": 22, "D": 28, "E": 22, "I": 40, "K": 26, "L": 26}.items(): ws.column_dimensions[col].width = w
    wb.save(os.path.join(A.dir, "leads-top-by-sponsor.xlsx"))
except ImportError: pass

# ---------- dashboard html ----------
n = len(REC); going = sum(1 for r in REC if r["status"] == "going")
counts = dict(mult=sum(1 for r in REC if r["multiplier"]), inv=sum(1 for r in REC if "Investor" in r["type"]),
              comp=sum(1 for r in REC if r["competitor_of"]), emp=sum(1 for r in REC if r["type"] == "Employee"))
SPHEAD = "".join(f'<th class="s" data-k="{s}">{s}</th>' for s in SP)
HTML = r"""<!doctype html><html><head><meta charset="utf-8"><title>Event Leads</title>
<meta name="viewport" content="width=device-width,initial-scale=1"><style>
:root{--navy:#__ACCENT__;--blue:#1E66B3;--bg:#f6f8fb;--line:#dde5ee}
*{box-sizing:border-box}body{font:14px/1.4 -apple-system,Segoe UI,Roboto,Arial;margin:0;background:var(--bg);color:#10233f}
header{background:#__ACCENT__;color:#fff;padding:14px 18px}header h1{margin:0 0 4px;font-size:18px}.sub{opacity:.85;font-size:12px}
.bar{position:sticky;top:0;z-index:5;background:#fff;border-bottom:1px solid var(--line);padding:10px 14px;display:flex;gap:8px;flex-wrap:wrap;align-items:center}
input#q{flex:1;min-width:220px;padding:8px 10px;border:1px solid var(--line);border-radius:8px;font-size:14px}
.chip{border:1px solid var(--line);background:#fff;border-radius:999px;padding:5px 11px;cursor:pointer;font-size:12px;user-select:none}
.chip.on{background:var(--blue);color:#fff;border-color:var(--blue)}
button.dl{border:1px solid var(--blue);color:var(--blue);background:#fff;border-radius:8px;padding:6px 10px;cursor:pointer;font-size:12px}
.count{font-size:12px;color:#5a6b82;margin-left:auto}
table{width:100%;border-collapse:collapse;background:#fff}th,td{padding:6px 9px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}
th{position:sticky;top:54px;background:#eef3f9;cursor:pointer;font-size:12px;white-space:nowrap}th.s,td.s{text-align:center}td.s{font-weight:700;color:#0a2a12}
tr.emp,tr.comp{background:#ffe3ea}tr.inv{background:#efe6ff}
.badge{display:inline-block;font-size:10px;padding:1px 6px;border-radius:6px;background:#eef3f9;color:#395a86;margin:1px 2px 1px 0}
.b-mult{background:#fff2cc;color:#7a5b00}.b-comp{background:#ffd6e0;color:#9c1f3a}.b-inv{background:#e6d6ff;color:#5a2ea6}.b-emp{background:#d6d6d6}
a{color:var(--blue);text-decoration:none}.note{font-size:11px;color:#7a8aa0;padding:8px 14px}
</style></head><body>
<header><h1>Event Leads — Intelligence</h1><div class="sub">__N__ engaged · __GO__ going · __MU__ multipliers · __IN__ investors · __CO__ competitors · __EM__ staff &nbsp;|&nbsp; rule-based + web-verified top tier · PII — keep local</div></header>
<div class="bar"><input id="q" placeholder="Search name, company, title, email, flags…">
<span class="chip on" data-f="all">All</span><span class="chip" data-f="going">Going</span><span class="chip" data-f="pending">Pending</span>
<span class="chip" data-f="mult">📣 Multipliers</span><span class="chip" data-f="inv">💰 Investors</span><span class="chip" data-f="comp">⚔️ Competitors</span>
<span class="chip" data-f="ver">✓ Verified</span><span class="chip" data-f="flag">⚠ Flagged</span><span class="chip" data-f="hi">📣 High-inf</span><span class="chip" data-f="vip">VIP</span><span class="chip" data-f="hot">Score ≥7</span>
<button class="dl" onclick="dl('hubspot')">⬇ HubSpot</button><button class="dl" onclick="dl('salesforce')">⬇ Salesforce</button><span class="count" id="cnt"></span></div>
<table id="t"><thead><tr><th data-k="rank">#</th><th data-k="name">Name</th><th data-k="company">Company</th><th data-k="title">Title</th><th data-k="type">Type</th>__SPHEAD__<th data-k="status">Status</th><th>Links</th></tr></thead><tbody id="tb"></tbody></table>
<div class="note">Scores 1–9 per sponsor from company+title. Sort by clicking a sponsor header. CRM export = current filtered view. PII — keep local.</div>
<script>
const D=__DATA__, SP=__SP__;let filt='all',q='',sortK='best',asc=false;
const sc=v=>{v=+v;return v>=8?'#1a7d3a':v>=6?'#3f9e57':v>=4?'#caa83a':v>=1?'#c0633a':'#999'};
function pass(r){
 if(filt==='going'&&r.status!=='going')return false; if(filt==='pending'&&r.status!=='pending')return false;
 if(filt==='mult'&&!r.multiplier)return false; if(filt==='inv'&&!/Investor/.test(r.type))return false;
 if(filt==='comp'&&!r.competitor_of)return false; if(filt==='vip'&&!r.vip)return false;
 if(filt==='ver'&&r.verified!=='V')return false; if(filt==='flag'&&!r.vflag)return false; if(filt==='hi'&&r.influence!=='high')return false;
 if(filt==='hot'&&Math.max(...SP.map(s=>+r[s]))<7)return false;
 if(q){const s=(r.name+' '+r.company+' '+r.title+' '+r.email+' '+r.flags+' '+r.type+' '+(r.notable||'')+' '+(r.vflag||'')+' '+(r.influence||'')).toLowerCase();if(!s.includes(q))return false}
 return true}
function rows(){let a=D.filter(pass);const num=[...SP,'best','rank'];
 a.sort((x,y)=>{let xv=num.includes(sortK)?+x[sortK]:(''+x[sortK]).toLowerCase(),yv=num.includes(sortK)?+y[sortK]:(''+y[sortK]).toLowerCase();return (xv<yv?-1:xv>yv?1:0)*(asc?1:-1)});return a}
function tag(r){let b='';
 if(r.verified)b+='<span class="badge" style="background:'+(r.verified=='V'?'#d8f5e0':r.verified=='P'?'#fff2cc':'#e6e6e6')+'">'+(r.verified=='V'?'✓ verified':r.verified=='P'?'~ partial':'? unverif')+'</span>';
 if(r.vflag)b+='<span class="badge b-comp">⚠ '+r.vflag+'</span>';
 if(r.multiplier)b+='<span class="badge b-mult">📣'+(r.influence?(' '+r.influence):'')+'</span>';
 if(r.competitor_of)b+='<span class="badge b-comp">⚔ '+r.competitor_of+'</span>';
 if(/Investor/.test(r.type))b+='<span class="badge b-inv">💰 investor</span>';
 if(r.type==='Employee')b+='<span class="badge b-emp">staff</span>';
 if(/bigco/.test(r.flags))b+='<span class="badge">bigco</span>';if(r.vip)b+='<span class="badge">VIP</span>';
 if(r.notable)b+='<div style="font-size:11px;color:#7a8aa0;margin-top:2px">'+r.notable+'</div>';return b}
function render(){const a=rows(),tb=document.getElementById('tb');
 tb.innerHTML=a.map(r=>{const cls=r.type==='Employee'?'emp':r.competitor_of?'comp':/Investor/.test(r.type)?'inv':'';
  const cells=SP.map(k=>`<td class="s" style="color:${sc(r[k])}">${r[k]}</td>`).join('');
  const li=r.linkedin?`<a href="${r.linkedin}" target="_blank">in</a>`:'';const em=r.email?` · <a href="mailto:${r.email}">✉</a>`:'';
  return `<tr class="${cls}"><td>${r.rank}</td><td><b>${r.name}</b><br>${tag(r)}</td><td>${r.company||''}</td><td>${r.title||''}</td><td><small>${r.type}</small></td>${cells}<td><small>${r.status}</small></td><td>${li}${em}</td></tr>`}).join('');
 document.getElementById('cnt').textContent=a.length+' / '+D.length+' shown';}
document.getElementById('q').oninput=e=>{q=e.target.value.toLowerCase();render()};
document.querySelectorAll('.chip').forEach(c=>c.onclick=()=>{document.querySelectorAll('.chip').forEach(x=>x.classList.remove('on'));c.classList.add('on');filt=c.dataset.f;render()});
document.querySelectorAll('th[data-k]').forEach(th=>th.onclick=()=>{const k=th.dataset.k;if(sortK===k)asc=!asc;else{sortK=k;asc=false}render()});
function csv(a,cols,map){const esc=s=>'"'+(''+(s??'')).replace(/"/g,'""')+'"';let o='﻿'+cols.join(',')+'\n';a.forEach(r=>o+=cols.map(c=>esc(map(r)[c])).join(',')+'\n');return o}
function dl(kind){const a=rows();const fn=s=>{const p=(s.name||'').trim().split(' ');return[p[0]||s.name,p.slice(1).join(' ')||'-']};let cols,map;
 if(kind==='hubspot'){cols=['First Name','Last Name','Email','Phone','Company Name','Job Title','LinkedIn','Lead Type','Multiplier',...SP.map(s=>s+' Score'),'Flags','Status'];
  map=r=>{const[f,l]=fn(r);const o={'First Name':f,'Last Name':l,'Email':r.email,'Phone':'','Company Name':r.company,'Job Title':r.title,'LinkedIn':r.linkedin,'Lead Type':r.type,'Multiplier':r.multiplier,'Flags':r.flags,'Status':r.status};SP.forEach(s=>o[s+' Score']=r[s]);return o}}
 else{cols=['First Name','Last Name','Company','Title','Email','Phone','Website','Lead Source','Rating','Description'];
  map=r=>{const[f,l]=fn(r);const best=Math.max(...SP.map(s=>+r[s]));const rate=best>=7?'Hot':best>=5?'Warm':'Cold';
   return{'First Name':f,'Last Name':l,'Company':r.company||'-','Title':r.title,'Email':r.email,'Phone':'','Website':r.linkedin,'Lead Source':'Event','Rating':rate,'Description':`Type:${r.type}; Mult:${r.multiplier}; Scores ${SP.map(s=>s[0]+r[s]).join('/')}; ${r.flags}`}}}
 const b=new Blob([csv(a,cols,map)],{type:'text/csv'});const u=URL.createObjectURL(b);const x=document.createElement('a');x.href=u;x.download=kind+'-leads-'+a.length+'.csv';x.click()}
render();
</script></body></html>"""
HTML = (HTML.replace("__DATA__", json.dumps(REC)).replace("__SP__", json.dumps(SP)).replace("__SPHEAD__", SPHEAD)
        .replace("__ACCENT__", ACCENT).replace("__N__", str(n)).replace("__GO__", str(going))
        .replace("__MU__", str(counts["mult"])).replace("__IN__", str(counts["inv"]))
        .replace("__CO__", str(counts["comp"])).replace("__EM__", str(counts["emp"])))
open(os.path.join(A.dir, "leads-dashboard.html"), "w").write(HTML)
print(f"wrote leads-dashboard.html ({len(HTML)//1024} KB), TOP-LEADS.md, leads-top-by-sponsor.xlsx  ({n} leads, {len(SP)} sponsors)")
