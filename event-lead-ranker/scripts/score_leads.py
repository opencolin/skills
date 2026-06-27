#!/usr/bin/env python3
"""Rule-based per-sponsor lead scoring from an event guest CSV. Config-driven.
Usage: python score_leads.py --config config.json [--csv guests.csv] [--out OUTDIR]
Scores ONLY engaged guests (config.engaged_statuses). Writes leads-ranked.{json,csv,xlsx}."""
import csv, json, collections, glob, os, argparse

ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True)
ap.add_argument("--csv", help="guest CSV; default = latest matching config.csv_glob")
ap.add_argument("--out", default=".", help="output directory")
A = ap.parse_args()
C = json.load(open(A.config))
SP = [s["name"] for s in C["sponsors"]]
src = A.csv or max(glob.glob(os.path.expanduser(C["csv_glob"])), key=os.path.getmtime)
cm = C["csv_columns"]
FREE = set(C["freemail"]); SPON_DOM = set(C["sponsor_domains"]); SPON_TOK = [t.lower() for t in C["sponsor_name_tokens"]]
VC = set(C["vc_domains"]); VCONF = C.get("vc_conflict", {})
COMP = C["competitors"]; BIGCO = C["bigco"]; FRONTIER = C["frontier_labs"]; DEVTOOL = C["devtool_ai"]
ENGAGED = set(C["engaged_statuses"]); SMAP = C.get("status_map", {})
MULT_T = C["multiplier_keywords"]["title"]; MULT_C = C["multiplier_keywords"]["company"]
BOOST_BIG = C.get("bigco_boost", {}); PEN_FRONT = C.get("frontier_penalty", {}); BOOST_DT = C.get("devtool_boost", [])

def dom(e): e = (e or "").lower(); return e.split("@")[1] if "@" in e else ""
def root(d): p = d.split("."); return ".".join(p[-2:]) if len(p) >= 2 else d
def has(h, ns): return any(n in h for n in ns)
def clamp(x): return max(1, min(9, int(round(x))))

rows = list(csv.DictReader(open(src, encoding="utf-8-sig")))
seen = {}
for r in rows:
    e = (r.get(cm["email"]) or "").strip().lower()
    if e and e not in seen: seen[e] = r
ded = [r for r in seen.values() if (r.get(cm["status"]) or "").strip() in ENGAGED]
print("SRC:", os.path.basename(src))
print("status distribution:", dict(collections.Counter((r.get(cm["status"]) or "").strip() for r in seen.values())))

def seniority(t):
    if has(t, ["founder","co-founder","cofounder","ceo","cto","chief","president"]): return 3, "founder/exec"
    if has(t, ["vp","vice president","head ","director","principal"]): return 2.5, "leader"
    if has(t, ["staff","lead "," lead","tech lead"]): return 2, "staff/lead"
    if has(t, ["senior","sr "]): return 1.5, "senior"
    if has(t, ["intern","student"]): return 0.3, "student/intern"
    if has(t, ["engineer","developer","scientist","researcher","manager","designer","architect"]): return 1.2, "ic"
    return 0.8, "other"

def fit(s, sp):
    f = sp["fit"]; return 3*has(s, f.get("3", [])) + 2*has(s, f.get("2", [])) + 1*has(s, f.get("1", []))

recs = []
for r in ded:
    name = (r.get(cm["name"]) or "").strip(); email = (r.get(cm["email"]) or "").strip().lower()
    d = dom(email); dr = root(d)
    company = (r.get(cm["company"]) or "").strip(); title = (r.get(cm["title"]) or "").strip()
    li = (r.get(cm.get("linkedin", "")) or "").strip(); status = (r.get(cm["status"]) or "").strip()
    vip = "vip" in (r.get(cm.get("ticket", "")) or "").lower()
    cl = company.lower(); tl = title.lower(); both = cl + " " + tl; s = tl + " " + cl
    is_edu = d.endswith(".edu") or d.endswith(".ac.uk") or has(cl, ["university","institute of technology","college","uc ","école"])
    flags = []; types = []
    sen, senlabel = seniority(tl)
    base = 2 + sen + (1 + (1 if vip else 0)) * 0.6
    scores = {sp["name"]: clamp(base + fit(s, sp) - (1.5 if is_edu else 0)) for sp in C["sponsors"]}
    if any(b in both for b in BIGCO):
        flags.append("bigco")
        for k, v in BOOST_BIG.items(): scores[k] = clamp(scores[k] + v)
    if any(fl in both for fl in FRONTIER):
        flags.append("frontier-lab")
        for k, v in PEN_FRONT.items(): scores[k] = clamp(scores[k] + v)
    if any(dt in both for dt in DEVTOOL):
        flags.append("ai-devtool")
        for k in BOOST_DT: scores[k] = clamp(scores[k] + 1)
    if is_edu: flags.append("student/edu")
    mult = has(tl, MULT_T) or has(cl, MULT_C)
    if mult: types.append("Multiplier"); flags.append("multiplier")
    is_vc = dr in VC or d in VC or has(cl, ["ventures","capital partners"," vc"]) or cl.endswith(" ventures") or cl.endswith(" capital")
    if is_vc:
        types.append("Investor"); flags.append("investor")
        for k in scores: scores[k] = 6
        for fund, ov in VCONF.items():
            if dr == fund or d == fund:
                for k, v in ov.items(): scores[k] = v
                flags.append(fund + "-conflict")
    comp_of = []
    for spn, c in COMP.items():
        if dr in set(c.get("domains", [])) or d in set(c.get("domains", [])) or has(both, c.get("names", [])):
            scores[spn] = 1; comp_of.append(spn)
    if comp_of: types.append("Competitor"); flags.append("competitor:" + ",".join(comp_of))
    is_emp = d in SPON_DOM or dr in SPON_DOM or any(tok == cl or tok in cl.split() for tok in SPON_TOK)
    if is_emp:
        types = ["Employee"]; flags = ["sponsor-staff:DO-NOT-PITCH"]
        for k in scores: scores[k] = 0
        comp_of = []; mult = False
    if not types:
        types.append("Founder" if senlabel == "founder/exec" else "Student" if senlabel == "student/intern"
                     else "Bigco" if any(b in both for b in BIGCO) else "Builder")
    rec = {"name": name, "company": company or (dr if dr and dr not in FREE else ""), "title": title,
           "type": "/".join(dict.fromkeys(types)), "multiplier": "YES" if mult else "", "competitor_of": ",".join(comp_of),
           "best": max(scores.values()), "status": SMAP.get(status, status), "email": email, "linkedin": li,
           "vip": "VIP" if vip else "", "flags": "; ".join(flags),
           "verified": "", "influence": "", "notable": "", "vflag": ""}
    rec.update(scores)
    recs.append(rec)

recs.sort(key=lambda x: (-x["best"], -sum(x[s] for s in SP), x["name"].lower()))
for i, r in enumerate(recs, 1): r["rank"] = i
cols = ["rank","name","company","title","type","multiplier","influence","verified","competitor_of"] + SP + \
       ["best","status","vip","notable","vflag","email","linkedin","flags"]
os.makedirs(A.out, exist_ok=True)
json.dump([{k: r.get(k, "") for k in cols} for r in recs], open(os.path.join(A.out, "leads-ranked.json"), "w"))
with open(os.path.join(A.out, "leads-ranked.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
    for r in recs: w.writerow({k: r.get(k, "") for k in cols})
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Leads"; ws.append(cols)
    for r in recs: ws.append([r.get(c, "") for c in cols])
    for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=C["ui"]["accent"])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; n = len(recs)
    for sp in SP:
        L = get_column_letter(cols.index(sp) + 1)
        ws.conditional_formatting.add(f"{L}2:{L}{n+1}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=5, mid_color="FFEB84", end_type="num", end_value=9, end_color="63BE7B"))
    pink = PatternFill("solid", fgColor="FFD6E0"); purple = PatternFill("solid", fgColor="E6D6FF")
    for i, r in enumerate(recs, 2):
        if r["type"] == "Employee" or r["competitor_of"]:
            for c in range(1, len(cols)+1): ws.cell(i, c).fill = pink
        elif "Investor" in r["type"]:
            for c in range(1, len(cols)+1): ws.cell(i, c).fill = purple
    wb.save(os.path.join(A.out, "leads-ranked.xlsx"))
except ImportError:
    print("(openpyxl missing — CSV/JSON written; `pip install openpyxl` for the formatted xlsx)")
n = len(recs)
print(f"SCORED {n} engaged  (going={sum(1 for r in recs if r['status']=='going')}, pending={sum(1 for r in recs if r['status']=='pending')})")
print(f"  Employees:{sum(1 for r in recs if r['type']=='Employee')} | Investors:{sum(1 for r in recs if 'Investor' in r['type'])} | Multipliers:{sum(1 for r in recs if r['multiplier'])} | Bigco:{sum(1 for r in recs if 'bigco' in r['flags'])}")
for sp in SP:
    print(f"  Competitors of {sp}: {sum(1 for r in recs if sp in r['competitor_of'])}")
print("-> leads-ranked.json / .csv / .xlsx in", os.path.abspath(A.out))
