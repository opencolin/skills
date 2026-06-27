#!/usr/bin/env python3
"""Merge deep-screen JSONL verification into leads-ranked.json: set verified/role/notable/flag/influence,
cap scores by flag, surface high-influence multipliers, re-rank, re-emit json/csv/xlsx.
Usage: python merge_screen.py --config config.json [--screen-dir /tmp] [--dir .]"""
import json, csv, glob, os, argparse, collections
ap = argparse.ArgumentParser()
ap.add_argument("--config", required=True)
ap.add_argument("--screen-dir", default="/tmp", help="dir with screen*.jsonl from the deep-screen agents")
ap.add_argument("--dir", default=".", help="dir holding leads-ranked.json")
A = ap.parse_args()
C = json.load(open(A.config)); SP = [s["name"] for s in C["sponsors"]]
path = os.path.join(A.dir, "leads-ranked.json"); REC = json.load(open(path))

def slug(u):
    u = (u or "").lower().rstrip("/")
    for m in ("/in/", "/company/"):
        if m in u: return u.split(m)[1].split("/")[0].split("?")[0]
    return ""

screen = {}
for fp in sorted(glob.glob(os.path.join(A.screen_dir, "screen*.jsonl"))):
    for line in open(fp):
        line = line.strip()
        if not line: continue
        try: o = json.loads(line)
        except Exception: continue
        if o.get("slug"): screen[o["slug"].lower()] = o
CAP = {"fake-company": 3, "micro-company": 4, "wrong-company": 4, "wrong-bigco": 5, "inflated-title": 6}
stats = collections.Counter(); applied = 0
for r in REC:
    s = slug(r.get("linkedin"))
    if s not in screen: continue
    o = screen[s]; applied += 1
    v = o.get("verdict", "").strip().upper()[:1]; flag = (o.get("flag", "") or "").strip()
    inf = (o.get("influence", "") or "").strip(); note = (o.get("notable", "") or "").strip()
    r["verified"] = v; stats[v] += 1
    if flag and flag != "none": r["vflag"] = flag; stats["flag:" + flag] += 1
    if note and note != "-": r["notable"] = note
    if inf and inf != "na": r["influence"] = inf
    if inf == "high": r["multiplier"] = "YES"
    cap = CAP.get(flag, 9)
    if v == "U": cap = min(cap, 4)
    for sp in SP: r[sp] = min(int(r[sp]), cap)
    r["best"] = max(int(r[sp]) for sp in SP)
REC.sort(key=lambda x: (-int(x["best"]), -sum(int(x[sp]) for sp in SP), x["name"].lower()))
for i, r in enumerate(REC, 1): r["rank"] = i
order = ["rank","name","company","title","type","multiplier","influence","verified","competitor_of"] + SP + \
        ["best","status","vip","notable","vflag","email","linkedin","flags"]
cols = [c for c in order if REC and c in REC[0]]
json.dump(REC, open(path, "w"))
with open(os.path.join(A.dir, "leads-ranked.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
    for r in REC: w.writerow({k: r.get(k, "") for k in cols})
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Leads"; ws.append(cols)
    for r in REC: ws.append([r.get(c, "") for c in cols])
    for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=C["ui"]["accent"])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; n = len(REC)
    for sp in SP:
        L = get_column_letter(cols.index(sp) + 1)
        ws.conditional_formatting.add(f"{L}2:{L}{n+1}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=5, mid_color="FFEB84", end_type="num", end_value=9, end_color="63BE7B"))
    pink = PatternFill("solid", fgColor="FFD6E0"); purple = PatternFill("solid", fgColor="E6D6FF")
    for i, r in enumerate(REC, 2):
        if r["type"] == "Employee" or r["competitor_of"]:
            for c in range(1, len(cols)+1): ws.cell(i, c).fill = pink
        elif "Investor" in r["type"]:
            for c in range(1, len(cols)+1): ws.cell(i, c).fill = purple
    wb.save(os.path.join(A.dir, "leads-ranked.xlsx"))
except ImportError: pass
print(f"applied screen to {applied} leads | verdicts {dict(stats)}")
for r in REC:
    if slug(r.get("linkedin")) in screen and r.get("vflag"):
        print(f"  ⚠ {r['vflag']:16} {r['name'][:26]:26} {r['company'][:22]}")
